# Copyright 2023-2024 David Goddard.
#
# This file is part of AwAg Data Services.
#
# AwAg Data Services is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License (AGPL), either
# version 3 of the License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful, but WITHOUT
# ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or
# FITNESS FOR A PARTICULAR PURPOSE. See the GNU Affero General Public License
# for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

import json
import io

from enum import Enum

from flask import current_app

import pandas as pd

from datetime import datetime, timedelta

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

from pylatex import Document, Ref, Label, Section, Subsection, Tabular, Command, Package, NoEscape, NewLine

from domestique.validation import Validator, NoiseLevel
from domestique.logging import log_exception

from .shared_resources import logger

from .awag_stats_engine_ext import StatsEngineExtended
from .awag_stats_builder_multidf import StatsBuilderMultiIndexDF


class StatsEngineExtendedDataFrames(StatsEngineExtended):

    def __init__(self, agent_id, conn, stats_table_base, stats_table_eval, flask_app=None, notes_dict={}):

        super().__init__(agent_id, conn, stats_table_base, stats_table_eval, flask_app, notes_dict)
        
        self.stats_builder_multidf = StatsBuilderMultiIndexDF(self, flask_app)


    def _create_df_from_stats_json(self, stats_json, index_label, data_columns):

        logger.debug(f"create_df_from_stats_json data_columns: {data_columns}")

        if not data_columns:
            raise ValueError("Bad (empty) data_columns")

        if not stats_json:
            logger.debug("Empty stats_json in _create_df_from_stats_json")
            return None

        overall_data = stats_json[self.LBL_ALL]
        if not overall_data:
            logger.debug("Empty overall_data in _create_df_from_stats_json")
            return None

        overall_df = pd.DataFrame([overall_data], index=[self.LBL_ALL])
        #logger.debug(f"Built overall_df:\n{overall_df}")

        classification_data = [metrics for metrics in stats_json[self.LBL_CLASS].values()]
        classification_df = pd.DataFrame(classification_data, index=stats_json[self.LBL_CLASS].keys())
        #logger.debug(f"Built classification_df:\n{classification_df}")

        combined_df = pd.concat([overall_df, classification_df])

        return_df = combined_df[data_columns]
        #logger.debug(f"Built DataFrame in create_df_from_stats_json:\n{return_df}")

        return return_df


    def _handle_time_series_data(self, stats_json):

        time_series_dfs = {}
        overall_data = stats_json[self.LBL_ALL]

        if "time_series" in overall_data:
            overall_time_series_df = pd.DataFrame(overall_data["time_series"]).set_index("date")
            time_series_dfs[self.LBL_ALL] = overall_time_series_df

        for classification, metrics in stats_json[self.LBL_CLASS].items():
            if "time_series" in metrics:
                classification_time_series_df = pd.DataFrame(metrics["time_series"]).set_index("date")
                logger.debug(f"Built time-series DataFrame:\n{classification_time_series_df}")
                time_series_dfs[classification] = classification_time_series_df

        return time_series_dfs


    def _convert_stats_df_pack_to_json(self, stats_df_pack_obj):

        if isinstance(stats_df_pack_obj, dict):
            return {k: self._convert_stats_df_pack_to_json(v) for k, v in stats_df_pack_obj.items()}
        elif isinstance(stats_df_pack_obj, pd.DataFrame):
            return self.df_to_json(stats_df_pack_obj)
        else:
            return stats_df_pack_obj


    def df_to_json(self, dataframe):

        if dataframe is None or dataframe.empty:
            return None
        else:
            df_json = dataframe.to_json(orient="index")
            return json.loads(df_json)


    def get_multiindex_df(self, stat_func, tag_main, tags_eval, **kwargs):

        return self.stats_builder_multidf.get_multiindex_df(stat_func, tag_main, tags_eval, **kwargs)


    def get_percent_of_classification_manual_agrees_base_df(self, tag_main, is_include_notes=True):

        stats_json = self.get_percent_of_classification_manual_agrees_base_cl(tag_main, is_include_notes)
        general_df = self._create_df_from_stats_json(stats_json, self.LBL_ALL, self.get_percent_of_classification_manual_agrees_base_df.data_columns)
        time_series_dfs = self._handle_time_series_data(stats_json)

        return general_df, time_series_dfs

    get_percent_of_classification_manual_agrees_base_df.data_columns = ["count_items", "count_agree", "percentage_agree", "dataset_size"]


    def get_pearsonr_for_eval_feedback_df(self, tag_main, tag_eval=None, persona_id=None, perspective_id=None, note_text=None, is_include_notes=True):

        stats_json = self.get_pearsonr_for_eval_feedback_cl(tag_main, tag_eval=tag_eval, persona_id=persona_id, perspective_id=perspective_id, note_text=note_text, is_include_notes=is_include_notes)

        general_df = self._create_df_from_stats_json(stats_json, self.LBL_ALL,
            self.get_pearsonr_for_eval_feedback_df.data_columns)

        return general_df

    get_pearsonr_for_eval_feedback_df.data_columns = ["correlation", "p_value", "item_count"]


    def get_pointbiserialr_for_eval_agreement_df(self, tag_main, tag_eval=None, persona_id=None, perspective_id=None, note_text=None, is_include_notes=True):

        stats_json = self.get_pointbiserialr_for_eval_agreement_cl(tag_main, tag_eval=tag_eval, persona_id=persona_id, perspective_id=perspective_id, note_text=note_text, is_include_notes=is_include_notes)

        general_df = self._create_df_from_stats_json(stats_json, self.LBL_ALL,
            self.get_pointbiserialr_for_eval_agreement_df.data_columns)

        return general_df

    get_pointbiserialr_for_eval_agreement_df.data_columns = ["correlation", "p_value", "item_count"]


    def get_pointbiserialr_for_eval_feedback_agreement_df(self, tag_main, tag_eval=None, persona_id=None, perspective_id=None, note_text=None, is_include_notes=True):

        stats_json = self.get_pointbiserialr_for_eval_feedback_agreement_cl(tag_main, tag_eval=tag_eval, persona_id=persona_id, perspective_id=perspective_id, note_text=note_text, is_include_notes=is_include_notes)

        general_df = self._create_df_from_stats_json(stats_json, self.LBL_ALL,
            self.get_pointbiserialr_for_eval_feedback_agreement_df.data_columns)

        return general_df

    get_pointbiserialr_for_eval_feedback_agreement_df.data_columns = ["correlation", "p_value", "item_count"]


    def get_avg_and_spread_for_eval_difference_df(self, tag_main, tag_eval=None, persona_id=None, perspective_id=None, note_text=None, is_include_notes=True):

        stats_json = self.get_avg_and_spread_for_eval_difference_cl(tag_main, tag_eval=tag_eval, persona_id=persona_id, perspective_id=perspective_id, note_text=note_text, is_include_notes=is_include_notes)

        general_df = self._create_df_from_stats_json(stats_json, self.LBL_ALL,
            self.get_avg_and_spread_for_eval_difference_df.data_columns)

        return general_df

    get_avg_and_spread_for_eval_difference_df.data_columns = ["average", "stddev", "median", "mode", "counts", "item_count"]


    def get_text_stats_df(self, tag_main, tag_eval=None, persona_id=None, perspective_id=None, note_text=None, is_include_notes=True, field_name=None):

        stats_json = self.get_text_stats_cl(tag_main, tag_eval=tag_eval, persona_id=persona_id, perspective_id=perspective_id, note_text=note_text, is_include_notes=is_include_notes,
        field_name=field_name)

        general_df = self._create_df_from_stats_json(stats_json, self.LBL_ALL,
            self.get_text_stats_df.data_columns)

        return general_df

    get_text_stats_df.data_columns = ["occurrences", "item_count"]


    def get_text_stats_df_dict(self, tag_main, tag_eval=None, persona_id=None, perspective_id=None, note_text=None, is_include_notes=True, field_names=None):

        stats_dfs = {}

        for field_name in field_names:
            stats_dfs[field_name] = self.get_text_stats_df(tag_main, tag_eval=tag_eval, persona_id=persona_id, perspective_id=perspective_id, note_text=note_text, is_include_notes=is_include_notes,
            field_name=field_name)

        return stats_dfs


    def get_text_stats_multidf_dict(self, tag_main, tags_eval=None, persona_id=None, perspective_id=None, note_text=None, is_include_notes=True, field_names=None):

        stats_dfs = {}

        if not field_names:
            return stats_dfs

        for field_name in field_names:
            stats_dfs[field_name] = self.get_multiindex_df(
            self.get_text_stats_df,
            tag_main, tags_eval, persona_id=persona_id, perspective_id=perspective_id,
            note_text=note_text, is_include_notes=is_include_notes,
            field_name=field_name)

        return stats_dfs


    def get_phi_coefficient_for_mode3_agreement_df(self, tag_main, tag_eval=None, persona_id=None, perspective_id=None, note_text=None, is_include_notes=True):

        stats_json = self.get_phi_coefficient_for_mode3_agreement_cl(tag_main, tag_eval=tag_eval, persona_id=persona_id, perspective_id=perspective_id, note_text=note_text, is_include_notes=is_include_notes)

        general_df = self._create_df_from_stats_json(stats_json, self.LBL_ALL,
            self.get_phi_coefficient_for_mode3_agreement_df.data_columns)

        return general_df

    get_phi_coefficient_for_mode3_agreement_df.data_columns = ["phi_coefficient", "p_value", "contingency_cells_desc", "contingency_cells_raw", "counts"]


    def get_cohens_kappa_for_manual_classification_agreement_df(self, tag_main, note_text=None, is_include_notes=True):

        stats_json = self.get_cohens_kappa_for_manual_classification_agreement_cl(tag_main, is_include_notes=is_include_notes)
        #logger.debug(f"Got stats_json: {stats_json}")

        general_df = self._create_df_from_stats_json(stats_json, self.LBL_ALL, self.get_cohens_kappa_for_manual_classification_agreement_df.data_columns)
        logger.debug(f"Got general_df: {general_df}")

        return general_df

    get_cohens_kappa_for_manual_classification_agreement_df.data_columns = ["kappa_score"]


    def build_stats_pack(self, tag_main, tags_eval, likert_fields, persona_id=None, perspective_id=None, include_empty=True, output_json=True):

        stats_functions = {
            "pearsonr_for_eval_feedback": self.get_pearsonr_for_eval_feedback_df,
            "pointbiserialr_for_eval_agreement": self.get_pointbiserialr_for_eval_agreement_df,
            "avg_and_spread_for_eval_difference": self.get_avg_and_spread_for_eval_difference_df,
            "phi_coefficient_for_mode3_agreement": self.get_phi_coefficient_for_mode3_agreement_df,
            "text_stats_evaluation_likert": self.get_text_stats_multidf_dict
        }

        stats_df_pack = {}
        stats_json_pack = {} if output_json else None

        # Get manual agrees base df separately due to different structure
        general_df, time_series_dfs = self.get_percent_of_classification_manual_agrees_base_df(tag_main)

        stats_df_pack["percent_of_classification_manual_agrees"] = {
            self.LBL_ALL: general_df,
            self.LBL_TS: time_series_dfs
        }

        # Get cohens_kappa_for_manual_classification_agreement_df separately also
        cohens_kappa_for_manual_classification_agreement_df = self.get_cohens_kappa_for_manual_classification_agreement_df(tag_main)
        logger.debug(f"Got stat 'cohens_kappa_for_manual_classification_agreement': {cohens_kappa_for_manual_classification_agreement_df}")

        stats_df_pack["cohens_kappa_for_manual_classification_agreement"] = cohens_kappa_for_manual_classification_agreement_df

        # Process others
        for key, func in stats_functions.items():
            if key == "text_stats_evaluation_likert":
                stats_df_pack[key] = func(tag_main, tags_eval, persona_id=persona_id, perspective_id=perspective_id, field_names=likert_fields)
            else:
                stats_df_pack[key] = self.get_multiindex_df(func, tag_main, tags_eval, persona_id=persona_id, perspective_id=perspective_id)

        if output_json:
            stats_json_pack = self._convert_stats_df_pack_to_json(stats_df_pack)

        return stats_df_pack, stats_json_pack


    def _write_dataframe_to_sheet(self, writer, sheet_name, df, full_name, index=True, start_row=0):

        df.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 2, index=index)
        worksheet = writer.sheets[sheet_name]

        worksheet.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+1, end_column=len(df.columns)+1)
        cell = worksheet.cell(row=start_row+1, column=1)
        cell.value = full_name
        cell.font = Font(bold=True, size=14)

        for col in range(1, len(df.columns) + 2):
            worksheet.cell(row=start_row+2, column=col).value = None


    def _write_dataframe_to_sheet(self, writer, sheet_name, df, df_title, index=True, gap=2):

        if sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            start_row = worksheet.max_row + gap
        else:
            worksheet = writer.book.create_sheet(sheet_name)
            start_row = 0

        df.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 2, index=index)
        worksheet = writer.sheets[sheet_name]

        worksheet.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=len(df.columns) + 1)
        cell = worksheet.cell(row=start_row + 1, column=1)
        cell.value = df_title
        cell.font = Font(bold=True, size=14)

        for col in range(1, len(df.columns) + 2):
            worksheet.cell(row=start_row + 2, column=col).value = None

        worksheet.column_dimensions['A'].width = 30
        for col in range(2, len(df.columns) + 2):
            worksheet.column_dimensions[chr(64 + col)].width = 18


    def _write_notes_to_sheet(self, writer, sheet_name, notes, gap=2):

        if not notes:
            return

        if sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            start_row = worksheet.max_row + gap
        else:
            worksheet = writer.book.create_sheet(sheet_name)
            start_row = 1

        for note_title, note_text in notes.items():
            worksheet.cell(row=start_row, column=1, value=note_title).font = Font(bold=True)
            worksheet.cell(row=start_row, column=2, value=note_text)
            start_row += 1


    def _preprocess_likert_df(self, df):

        dataframes = {}

        logger.debug(f"_preprocess_likert_df: {df}")

        for tag_eval in df.columns.get_level_values(0).unique():
            logger.debug(f"tag_eval: {tag_eval}")
            if tag_eval == 'metric':
                continue

            col_df = df[tag_eval].reset_index()
            col_df.columns = ["metric"] + list(col_df.columns[1:])
            dataframes[tag_eval] = col_df

        return dataframes


    def _expand_occurrences_column(self, df):

        rows = []

        for _, row in df.iterrows():
            base_data = row.drop("occurrences").to_dict() if "occurrences" in row else row.to_dict()
            occurrences = row["occurrences"]

            for key, value in occurrences.items():
                new_row = base_data.copy()
                new_row.update({
                    "response": key,
                    "count": value["count"],
                    "percentage": value["percentage"]
                })
                rows.append(new_row)

        expanded_df = pd.DataFrame(rows)
        expanded_df.set_index(["metric", "response"], inplace=True)

        logger.debug(f"_expand_occurrences_column returning: {expanded_df}")

        return expanded_df


    def convert_stats_pack_to_excel(self, stats_df_pack, engine="openpyxl", index=True, tag_main=""):

        """
        Convert the stats_df_pack into a Excel format.
        """

        output = io.BytesIO()

        def get_short_key_name(key):
            return '_'.join([part[0] for part in key.split('_')])

        with pd.ExcelWriter(output, engine=engine) as writer:

            for key, value in stats_df_pack.items():

                sheet_name_short = get_short_key_name(key)
                df_title_base = ' '.join([word.capitalize() for word in key.split('_')])

                if isinstance(value, pd.DataFrame):

                    self._write_dataframe_to_sheet(writer, sheet_name_short, value, df_title_base, index)

                elif isinstance(value, dict):

                    if key == "percent_of_classification_manual_agrees":

                        self._write_dataframe_to_sheet(writer, f"{sheet_name_short}_all", value["ALL"], f"{df_title_base} [all]", index)

                        sheet_name_ts = f"{sheet_name_short}_ts"
                        for sub_key, df in value["TIMESERIES"].items():
                            df_title = f"{df_title_base} - time series for: {sub_key}"
                            self._write_dataframe_to_sheet(writer, sheet_name_ts, df, df_title, index)

                    elif key == "text_stats_evaluation_likert":

                        for sub_key, df in value.items():

                            sub_key_short = get_short_key_name(key)
                            sheet_name = f"{sheet_name_short}_{sub_key_short}"

                            logger.debug(f"Outputting sub_key '{sub_key}' - sheet_name: {sheet_name}")

                            preprocessed_dfs = self._preprocess_likert_df(df)

                            for tag_eval, preprocessed_df in preprocessed_dfs.items():
                                logger.debug(f"Outputting tag_eval '{tag_eval}' for sub_key '{sub_key}' - sheet_name: {sheet_name}")
                                df_title = f"Likert Stats: {tag_eval} - {sub_key}"
                                expanded_df = self._expand_occurrences_column(preprocessed_df)
                                self._write_dataframe_to_sheet(writer, sheet_name, expanded_df, df_title, index)

            for key in stats_df_pack.keys():
                if key == "percent_of_classification_manual_agrees":
                    sheet_name_short = f"{get_short_key_name(key)}_all"
                else:
                    sheet_name_short = get_short_key_name(key)
                notes = self.get_notes(key, tag_main=tag_main) 
                self._write_notes_to_sheet(writer, sheet_name_short, notes)

        output.seek(0)

        return output.getvalue()


    def convert_stats_pack_to_latex(self, agent_name, stats_df_pack, tag_main="", minimal=False):

        """
        Convert the stats_df_pack into a LaTeX document.
        """

        doc = Document()
        doc.packages.append(Package('geometry', options=['a4paper']))

        if minimal:
            doc.append(f"MINIMAL")
        else:
            title = f"Stats Report for \\texttt{{{agent_name}}} / \\texttt{{{tag_main}}}"
            doc.preamble.append(Command("title", NoEscape(title)))
            doc.preamble.append(Command("author", "AwAg Data Services"))
            doc.preamble.append(Command("date", NoEscape("\\today")))
            doc.append(NoEscape("\\maketitle"))

        # # Generate sections and tables
        # for key, value in stats_df_pack.items():

            # section_title = key.replace('_', ' ').title()

            # with doc.create(Section(section_title)):
                # if isinstance(value, pd.DataFrame):
                    # with doc.create(Subsection("Overall")):
                        # with doc.create(Tabular('|'.join(['l'] * (len(value.columns) + 1)))) as table:
                            # table.add_hline()
                            # table.add_row(["Metric"] + list(value.columns))
                            # table.add_hline()
                            # for index, row in value.iterrows():
                                # table.add_row([index] + row.tolist())
                            # table.add_hline()
                    # #etc.


        def get_initials(key):

            return '-'.join([part[0] for part in key.split('_')])

        def get_id_from_key(key):

            return key.replace('_', '-')

        def get_title_from_key(key):

            return key.replace('_', ' ').title()

        def get_section_id(key):

            return f"stats-{agent_name}-{get_initials(key)}"

        def get_section_title(key):

            return f"{agent_name} -- {get_title_from_key(key)}"


        def get_table_caption(title):

            return f"{title} ({agent_name.title()})"

        def get_label(prefix, id):

            return NoEscape(f"{prefix}:{id}")

        def get_labelX(prefix, id):

            return f"{prefix}:{id}"

        def append_label(label):

            # Workaround for pylatex always escaping hyphens in label IDs no matter what I try...
            doc.append(Label(NoEscape(label)))


        def get_table_identifiers(id, title):
        
            table_label = get_label("tab", id)
            table_caption = get_table_caption(title)
            return table_label, table_caption


        for key, value in stats_df_pack.items():

            section_id = get_section_id(key)
            section_title = NoEscape(get_section_title(key))  # Same as df_title_base
            section_label = get_label("sec", section_id)

            
            #with doc.create(Section(section_title, label=section_label)):
            #with doc.create(Section(section_title, label=NoEscape(section_id))):
            with doc.create(Section(section_title, label=False)):
                # It appears to be impossible to pass a label param to Section and have this used unescaped, so DIY...
                #append_label(
                doc.append(Label(section_label))
                doc.append(Label(NoEscape(get_labelX("sec", section_id))))
                timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                doc.append(section_label)

                if isinstance(value, pd.DataFrame):

                    # This section contains only one table

                    table_label, table_caption = get_table_identifiers(section_id, section_title)

                    doc.append(f"Table ")
                    doc.append(Ref(table_label))
                    doc.append(f": {table_caption}")

                    # TODO: Output table using label and caption

                elif isinstance(value, dict):

                    if key == "percent_of_classification_manual_agrees":

                        subsection_title = f"{section_title} - time series: ALL"
                        subsection_id = f"{section_id}-ALL"
                        subsection_label = get_label("subsec", subsection_id)

                        with doc.create(Subsection(section_title)):

                            doc.append(Label(subsection_label))

                            table_label, table_caption = get_table_identifiers(subsection_id, subsection_title)

                            doc.append(f"Table ")
                            doc.append(Ref(table_label))
                            doc.append(f": {table_caption}")

                            # TODO: Output table using label and caption


                        subsection_id_base = f"{section_id}-ts-"

                        #sheet_name_ts = f"{sheet_name_short}_ts"
                        for sub_key, df in value["TIMESERIES"].items():

                            sub_id = get_initials(sub_key)
                            subsection_id = f"{section_id}-ts-{sub_id}"
                            subsection_title = f"{section_title} - time series: {sub_key}"
                            subsection_label = get_label("subsec", subsection_id)

                            with doc.create(Subsection(subsection_title)):

                                doc.append(Label(subsection_label))

                                table_label, table_caption = get_table_identifiers(subsection_id, subsection_title)

                                doc.append(f"Table ")
                                doc.append(Ref(table_label))
                                doc.append(f": {table_caption}")

                                # TODO: Output table using label and caption


                    elif key == "text_stats_evaluation_likert":

                        for sub_key, df in value.items():

                            subsection_title = f"{section_title} - {sub_key}"
                            subsection_id = get_id_from_key(sub_key)
                            subsection_label = get_label("subsec", subsection_id)


                            with doc.create(Subsection(subsection_title)):

                                doc.append(Label(subsection_label))

                                doc.append(f"sub_key: {sub_key}")
                                doc.append(NewLine())
                                doc.append(f"section_id: {section_id}")
                                doc.append(NewLine())
                                doc.append(f"subsection_id: {subsection_id}")
                                doc.append(NewLine())

                                # Need to do more work before we can output tables for this

        full_output = doc.dumps()

        if minimal:
            start_marker = "\\begin{document}"
            end_marker = "\\end{document}"
            start_idx = full_output.index(start_marker) + len(start_marker)
            end_idx = full_output.index(end_marker)
            return full_output[start_idx:end_idx].strip()  # Return only the content between \begin and \end
        else:
            return full_output
